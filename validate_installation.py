#!/usr/bin/env python3
"""Validation script to ensure ETL Dashboard is properly installed and configured."""

import sys
import os
import importlib
from pathlib import Path
import pandas as pd
from datetime import datetime

def check_python_version():
    """Check Python version compatibility."""
    print("🐍 Checking Python version...")
    version = sys.version_info
    if version >= (3, 11):
        print(f"✅ Python {version.major}.{version.minor}.{version.micro} is compatible")
        return True
    else:
        print(f"❌ Python {version.major}.{version.minor}.{version.micro} is too old. Requires 3.11+")
        return False

def check_dependencies():
    """Check if all required dependencies are installed."""
    print("📦 Checking dependencies...")
    
    required_packages = [
        'fastapi', 'uvicorn', 'flask', 'pandas', 'numpy', 
        'openpyxl', 'pyarrow', 'fastparquet', 'pydantic',
        'sqlalchemy', 'structlog', 'python_dotenv'
    ]
    
    missing_packages = []
    
    for package in required_packages:
        try:
            # Handle special cases
            if package == 'python_dotenv':
                importlib.import_module('dotenv')
            else:
                importlib.import_module(package)
            print(f"   ✅ {package}")
        except ImportError:
            print(f"   ❌ {package}")
            missing_packages.append(package)
    
    if missing_packages:
        print(f"❌ Missing packages: {', '.join(missing_packages)}")
        return False
    else:
        print("✅ All dependencies are installed")
        return True

def check_directory_structure():
    """Check if required directories exist."""
    print("📁 Checking directory structure...")
    
    required_dirs = [
        'backend', 'backend/api', 'backend/core', 'backend/models', 'backend/services',
        'frontend', 'frontend/templates', 'frontend/static', 'frontend/static/css', 'frontend/static/js',
        'data', 'data/uploads', 'data/processed',
        'powerbi', 'tests'
    ]
    
    missing_dirs = []
    
    for directory in required_dirs:
        if Path(directory).exists():
            print(f"   ✅ {directory}")
        else:
            print(f"   ❌ {directory}")
            missing_dirs.append(directory)
    
    if missing_dirs:
        print(f"❌ Missing directories: {', '.join(missing_dirs)}")
        return False
    else:
        print("✅ All required directories exist")
        return True

def check_core_files():
    """Check if core application files exist."""
    print("📄 Checking core files...")
    
    required_files = [
        'backend/main.py',
        'backend/core/config.py',
        'backend/core/logging.py',
        'backend/models/schemas.py',
        'backend/services/excel_reader.py',
        'backend/services/cleaning.py',
        'backend/services/masterbom_rules.py',
        'backend/api/routes_upload.py',
        'frontend/app.py',
        'frontend/templates/base.html',
        'frontend/templates/index.html',
        'frontend/static/js/app.js',
        'requirements.txt',
        'README.md'
    ]
    
    missing_files = []
    
    for file_path in required_files:
        if Path(file_path).exists():
            print(f"   ✅ {file_path}")
        else:
            print(f"   ❌ {file_path}")
            missing_files.append(file_path)
    
    if missing_files:
        print(f"❌ Missing files: {', '.join(missing_files)}")
        return False
    else:
        print("✅ All core files exist")
        return True

def test_core_functionality():
    """Test core functionality."""
    print("🧪 Testing core functionality...")
    
    try:
        # Test data cleaning
        from backend.services.cleaning import clean_id
        test_id = clean_id("  7009@6933#  ")
        assert test_id == "70096933", f"Expected '70096933', got '{test_id}'"
        print("   ✅ ID cleaning works")
        
        # Test date parsing
        from backend.services.cleaning import parse_date_column
        test_dates = pd.Series(["2024-01-15", "2024-02-20"])
        result = parse_date_column(test_dates, "test_date")
        assert "test_date_year" in result.columns, "Date parsing failed"
        print("   ✅ Date parsing works")
        
        # Test profiler
        from backend.services.profiler import DataProfiler
        test_df = pd.DataFrame({"col1": [1, 2, 3], "col2": ["A", "B", "C"]})
        profiler = DataProfiler(test_df, "test_sheet")
        profile = profiler.profile_sheet()
        assert profile.total_rows == 3, "Profiler failed"
        print("   ✅ Data profiler works")
        
        print("✅ Core functionality tests passed")
        return True
        
    except Exception as e:
        print(f"❌ Core functionality test failed: {e}")
        return False

def check_configuration():
    """Check configuration files."""
    print("⚙️  Checking configuration...")
    
    # Check .env file
    env_file = Path(".env")
    env_example = Path(".env.example")
    
    if env_example.exists():
        print("   ✅ .env.example exists")
    else:
        print("   ❌ .env.example missing")
        return False
    
    if env_file.exists():
        print("   ✅ .env exists")
    else:
        print("   ⚠️  .env not found (will use defaults)")
    
    # Test configuration loading
    try:
        from backend.core.config import settings
        print(f"   ✅ Configuration loaded (upload folder: {settings.upload_folder})")
        return True
    except Exception as e:
        print(f"   ❌ Configuration loading failed: {e}")
        return False

def create_sample_data():
    """Create sample data for testing."""
    print("📊 Creating sample test data...")
    
    try:
        # Create sample Excel file
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        ws.title = "MasterBOM"
        
        # Headers
        headers = ["YAZAKI PN", "PROJECT_1", "PROJECT_2", "Item Description", "Supplier Name"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Sample data
        data = [
            ["7009-6933", "X", "", "Wire Harness", "Yazaki Corp"],
            ["7116-4101", "D", "X", "Connector", "Molex Inc"],
            ["ABC-123", "", "0", "Terminal", "TE Connectivity"]
        ]
        
        for row, row_data in enumerate(data, 2):
            for col, value in enumerate(row_data, 1):
                ws.cell(row=row, column=col, value=value)
        
        # Add Status sheet
        ws_status = wb.create_sheet("Status")
        ws_status.cell(row=1, column=1, value="OEM")
        ws_status.cell(row=1, column=2, value="Project")
        ws_status.cell(row=2, column=1, value="Toyota")
        ws_status.cell(row=2, column=2, value="Test Project")
        
        sample_file = Path("data/uploads/sample_test.xlsx")
        wb.save(sample_file)
        
        print(f"   ✅ Sample Excel file created: {sample_file}")
        return True
        
    except Exception as e:
        print(f"   ❌ Failed to create sample data: {e}")
        return False

def run_validation():
    """Run all validation checks."""
    print("🔍 ETL Dashboard Installation Validation")
    print("=" * 45)
    print(f"Timestamp: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print()
    
    checks = [
        ("Python Version", check_python_version),
        ("Dependencies", check_dependencies),
        ("Directory Structure", check_directory_structure),
        ("Core Files", check_core_files),
        ("Configuration", check_configuration),
        ("Core Functionality", test_core_functionality),
        ("Sample Data", create_sample_data)
    ]
    
    passed = 0
    failed = 0
    
    for check_name, check_func in checks:
        print(f"\n{check_name}:")
        print("-" * len(check_name))
        
        if check_func():
            passed += 1
        else:
            failed += 1
    
    print("\n" + "=" * 45)
    print("📊 VALIDATION SUMMARY")
    print("=" * 45)
    print(f"✅ Passed: {passed}")
    print(f"❌ Failed: {failed}")
    print(f"📈 Success Rate: {(passed/(passed+failed)*100):.1f}%")
    
    if failed == 0:
        print("\n🎉 All validation checks passed!")
        print("Your ETL Dashboard installation is ready to use.")
        print("\n🚀 Next steps:")
        print("1. Run: python run_dev.py")
        print("2. Open: http://localhost:5000")
    else:
        print(f"\n⚠️  {failed} validation check(s) failed.")
        print("Please review the errors above and fix them before proceeding.")
        return False
    
    return True

if __name__ == "__main__":
    success = run_validation()
    sys.exit(0 if success else 1)
